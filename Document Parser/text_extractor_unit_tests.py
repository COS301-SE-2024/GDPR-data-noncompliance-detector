import unittest
from text_extractor import text_extractor
import tempfile
import os
from openpyxl import Workbook
from docx import Document
from reportlab.pdfgen import canvas
import pandas as pd

class TestTextExtractor(unittest.TestCase):
    def setUp(self):
        self.extractor = text_extractor()

    def test_extract_text_from_pdf(self):
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as temp:
            c = canvas.Canvas(temp.name)
            c.drawString(100, 750, "Test text")
            c.save()

            result = self.extractor.extract_text_from_pdf(temp.name)            
            self.assertEqual(result, 'Test text\n\n\x0c')
        os.remove(temp.name)

    def test_extract_text_from_docx(self):
        doc = Document()
        doc.add_paragraph('Test text')
        
        with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as temp:
            doc.save(temp.name)
            result = self.extractor.extract_text_from_docx(temp.name)
            self.assertEqual(result, 'Test text')
        
        os.remove(temp.name)

    def test_extract_data_from_excel(self):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Test text'
        
        with tempfile.NamedTemporaryFile(suffix=".xls", delete=False) as temp:
            wb.save(temp.name)
            result = self.extractor.extract_data_from_excel(temp.name)
            expected_result = 'Test text'
            self.assertEqual(result, expected_result)
        
        os.remove(temp.name)

    def test_extract_text(self):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Test text'
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp:
            wb.save(temp.name)
            result = self.extractor.extract_data_from_excel(temp.name)
            expected_result = 'Test text'
            self.assertEqual(result, expected_result)
        
        os.remove(temp.name)

if __name__ == '__main__':
    unittest.main()