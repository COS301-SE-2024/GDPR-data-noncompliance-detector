import pandas as pd
from pdfminer.high_level import extract_text
from docx import Document
from docx.shared import Inches
import pytesseract
from PIL import Image
import io
import pdfplumber
import openpyxl #we need this for excel
from openpyxl_image_loader import SheetImageLoader
from pdf2image import convert_from_path
class text_extractor:
    def __init__(self):
        self.ext = ''

    def extract_text_from_pdf(self, file_path):
        combined_text = []

        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    combined_text.append(text)

        images = convert_from_path(file_path)

        for image in images:
            ocr_text = pytesseract.image_to_string(image)
            combined_text.append(ocr_text)

        return "\n".join(combined_text)


    def extract_text_from_docx(self, file_path):
        document = Document(file_path)
        combined_text = []

        for para in document.paragraphs:
            combined_text.append(para.text)
        
        for rel in document.part.rels:
            if "image" in document.part.rels[rel].target_ref:
                img = document.part.rels[rel].target_part.blob
                image = Image.open(io.BytesIO(img))
                
                ocr_text = pytesseract.image_to_string(image)
                combined_text.append(ocr_text)
        
        return "\n".join(combined_text)

    def extract_data_from_excel(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        combined_text = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            df = pd.read_excel(file_path, sheet_name=sheet_name)
            if not df.empty:
                combined_text.append(df.to_string(index=False))

            try:
                image_loader = SheetImageLoader(sheet)
                for row in sheet.iter_rows():
                    for cell in row:
                        if image_loader.image_in(cell.coordinate):
                            image = image_loader.get(cell.coordinate)
                            ocr_text = pytesseract.image_to_string(image)
                            combined_text.append(ocr_text)
            except Exception as e:
                    print(f"Error processing image: {e}")

        return "\n".join(combined_text)

    def extract_text_multi(self, file_path, extension):
        if extension == '.pdf':
            text = self.extract_text_from_pdf(file_path)
        elif extension == '.docx':
            text = self.extract_text_from_docx(file_path)
        elif extension in ['.xlsx', '.xls']:
            text = self.extract_data_from_excel(file_path)
        else:
            text = None

        return text
