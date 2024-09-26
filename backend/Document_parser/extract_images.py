from pypdf import PdfReader
import os, sys
import docx2txt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

class image_extractor:
    def __init__(self):
        self.base_dir = os.path.dirname(__file__)

    def extract_images_from_excel(self, excel_file):
        # output_folder = os.path.join(self.base_dir, "../Detection_Engine/extracted_images/xlsx_images")
        output_folder = resource_path("../Detection_Engine/extracted_images/xlsx_images")
        
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
        
        workbook = load_workbook(excel_file)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            for image in sheet._images:
                img_format = 'png'
                if image.path.lower().endswith('.jpeg') or image.path.lower().endswith('.jpg'):
                    img_format = 'jpeg'
                elif image.path.lower().endswith('.bmp'):
                    img_format = 'bmp'
                
                img_filename = f"{sheet_name}_{image.anchor._from.row}_{image.anchor._from.col}.{img_format.lower()}"
                img_path = os.path.join(output_folder, img_filename)
                
                with open(img_path, 'wb') as f:
                    f.write(image._data())

    def extract_images_from_pdf(self, dir):
        # output_folder = os.path.join(self.base_dir, "../Detection_Engine/extracted_images/pdf_images")
        output_folder = resource_path("../Detection_Engine/extracted_images/pdf_images")
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        pypdf_reader = PdfReader(dir)

        for page in pypdf_reader.pages:
            for image in page.images:
                image_path = os.path.join(output_folder, image.name)
                with open(image_path, "wb") as fp:
                    fp.write(image.data)

    def extract_images_from_docx(self, docx_path):
        # output_folder = os.path.join(self.base_dir, "../Detection_Engine/extracted_images/docx_images")
        output_folder = resource_path("../Detection_Engine/extracted_images/docx_images")
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        docx2txt.process(docx_path, output_folder)
    # def extract_images_from_excel(self, excel_file):
    #     output_folder = "./Detection_Engine/extracted_images/xlsx_images"
        
    #     if not os.path.exists(output_folder):
    #         os.makedirs(output_folder)
        
    #     workbook = load_workbook(excel_file)
        
    #     for sheet_name in workbook.sheetnames:
    #         sheet = workbook[sheet_name]
            
    #         for image in sheet._images:
    #             img_format = 'png'
    #             if image.path.lower().endswith('.jpeg') or image.path.lower().endswith('.jpg'):
    #                 img_format = 'jpeg'
    #             elif image.path.lower().endswith('.bmp'):
    #                 img_format = 'bmp'
                
    #             img_filename = f"{sheet_name}_{image.anchor._from.row}_{image.anchor._from.col}.{img_format.lower()}"
    #             img_path = os.path.join(output_folder, img_filename)
                
    #             with open(img_path, 'wb') as f:
    #                 f.write(image._data())
                


    # def extract_images_from_pdf(self, dir):
    #     output_folder = "./Detection_Engine/extracted_images/pdf_images"
    #     if not os.path.exists(output_folder):
    #         os.makedirs(output_folder)

    #     # ensure folders are removed/replaced after each run
    #     # add to document parser
    #     pypdf_reader = PdfReader(dir)

    #     # this folder is made detection engine extracted images
    #     output_dir = "./Detection_Engine/extracted_images/pdf_images"
    #     os.makedirs(output_dir, exist_ok=True)

    #     for page in pypdf_reader.pages:
    #         for image in page.images:
    #             image_path = os.path.join(output_dir, image.name)
    #             with open(image_path, "wb") as fp:
    #                 fp.write(image.data)


    # def extract_images_from_docx(self, docx_path):
    #     output_folder = "./Detection_Engine/extracted_images/docx_images"
    #     if not os.path.exists(output_folder):
    #         os.makedirs(output_folder)

    #     docx2txt.process(docx_path, '../Detection_Engine/extracted_images/docx_images')
